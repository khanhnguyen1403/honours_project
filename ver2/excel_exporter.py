import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


class ExcelExporter:
    """
    Simplified Excel export utility focused on power consumption data.
    Creates concise Excel reports with essential power data only.
    """
    
    def __init__(self, appliances, right_gui):
        """Initialize the Excel exporter."""
        self.appliances = appliances
        self.right_gui = right_gui
        self.export_folder = "exports"
        
        # Create exports directory if it doesn't exist
        if not os.path.exists(self.export_folder):
            os.makedirs(self.export_folder)
    
    def export_data(self):
        """Export power consumption data to Excel file."""
        try:
            # Generate filename with timestamp
            timestamp = datetime.now()
            filename = f"appliance_data_{timestamp.strftime('%Y%m%d_%H%M')}.xlsx"
            filepath = os.path.join(self.export_folder, filename)
            
            # Create workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Power Data"
            
            # Create the simplified report
            self._create_power_report(sheet, timestamp)
            
            # Save file
            workbook.save(filepath)
            
            # Log success
            if hasattr(self.right_gui, 'log_events'):
                self.right_gui.log_events(f"Power data exported to {filename}")
            
            print(f"Excel file exported: {filepath}")
            return True
            
        except Exception as e:
            # Log error
            error_msg = f"Export failed: {str(e)}"
            if hasattr(self.right_gui, 'log_events'):
                self.right_gui.log_events(error_msg)
            print(error_msg)
            return False
    
    def _create_power_report(self, sheet, timestamp):
        """Create a concise power consumption report."""
        # Setup styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        
        # Title and timestamp
        sheet['A1'] = "Power Consumption Report"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet.merge_cells('A1:E1')
        
        sheet['A2'] = f"Export Time: {timestamp.strftime('%Y-%m-%d %H:%M:%S')}"
        sheet['A2'].font = Font(bold=True)
        sheet.merge_cells('A2:E2')
        
        # System totals section
        self._add_system_totals(sheet)
        
        # Individual appliances section
        self._add_individual_appliances(sheet, header_font, header_fill, center_align)
        
        # Power history section (last 10 readings for brevity)
        self._add_recent_power_history(sheet, header_font, header_fill, center_align)
        
        # Auto-adjust column widths
        self._auto_adjust_columns(sheet)
    
    def _add_system_totals(self, sheet):
        """Add system-wide power totals."""
        sheet['A4'] = "System Totals"
        sheet['A4'].font = Font(bold=True, size=14)
        
        if "All" in self.appliances and self.appliances["All"]:
            summary = self.appliances["All"]
            total_consumption = getattr(summary, 'total_power_consumption', 0)
            total_generation = getattr(summary, 'total_power_generation', 0)
            net_power = total_consumption - total_generation
            
            sheet['A5'] = "Total Power Consumption:"
            sheet['B5'] = f"{total_consumption:.1f} W"
            sheet['A6'] = "Total Power Generation:"
            sheet['B6'] = f"{total_generation:.1f} W"
            sheet['A7'] = "Net Power:"
            sheet['B7'] = f"{net_power:.1f} W"
            
            # Bold labels
            for row in [5, 6, 7]:
                sheet[f'A{row}'].font = Font(bold=True)
    
    def _add_individual_appliances(self, sheet, header_font, header_fill, center_align):
        """Add individual appliance power data."""
        start_row = 9
        sheet[f'A{start_row}'] = "Individual Appliances"
        sheet[f'A{start_row}'].font = Font(bold=True, size=14)
        
        # Headers
        headers = ["Appliance", "Status", "Current Power (W)", "Energy Used (kWh)"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row + 2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        # Data rows
        row = start_row + 3
        for name, appliance in self.appliances.items():
            if name == "All" or appliance is None:
                continue
                
            try:
                # Get power data safely
                status = "ON" if getattr(appliance, 'power_status', False) else "OFF"
                current_power = self._safe_get_power(appliance)
                energy_used = getattr(appliance, 'energy_used', 0)
                
                # Add to sheet
                sheet.cell(row=row, column=1, value=name)
                sheet.cell(row=row, column=2, value=status)
                sheet.cell(row=row, column=3, value=f"{current_power:.1f}")
                sheet.cell(row=row, column=4, value=f"{energy_used:.3f}")
                
                # Color code status
                status_cell = sheet.cell(row=row, column=2)
                if status == "ON":
                    status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                row += 1
                
            except Exception as e:
                print(f"Error processing appliance {name}: {e}")
    
    def _add_recent_power_history(self, sheet, header_font, header_fill, center_align):
        """Add recent power history (last 10 readings) for each appliance."""
        # Calculate starting row
        appliance_count = len([a for name, a in self.appliances.items() if name != "All" and a is not None])
        start_row = 12 + appliance_count + 2
        
        sheet[f'A{start_row}'] = "Recent Power History (Last 10 Readings)"
        sheet[f'A{start_row}'].font = Font(bold=True, size=14)
        
        # Create headers with appliance names
        headers = ["Time Index"] + [name for name, a in self.appliances.items() if name != "All" and a is not None]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=start_row + 2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        # Add last 10 power readings
        for i in range(10):
            row = start_row + 3 + i
            sheet.cell(row=row, column=1, value=f"T-{9-i}")  # Time index (T-9, T-8, ..., T-0)
            
            col = 2
            for name, appliance in self.appliances.items():
                if name == "All" or appliance is None:
                    continue
                    
                # Get power history safely
                power_history = self._safe_get_history(appliance)
                if power_history and len(power_history) > (9-i):
                    power_value = power_history[-(10-i)]  # Get value from end of array
                else:
                    power_value = 0
                    
                sheet.cell(row=row, column=col, value=f"{power_value:.1f}")
                col += 1
    
    def _safe_get_power(self, appliance):
        """Safely get current power from appliance."""
        try:
            if hasattr(appliance, 'get_current_power'):
                return appliance.get_current_power()
            return 0
        except:
            return 0
    
    def _safe_get_history(self, appliance):
        """Safely get power history from appliance."""
        try:
            if hasattr(appliance, 'get_power_history'):
                return appliance.get_power_history()
            return [0] * 300
        except:
            return [0] * 300
    
    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths."""
        try:
            # Get the actual range of columns used
            max_col = sheet.max_column
            
            for col_num in range(1, max_col + 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                
                # Check all cells in this column
                for row_num in range(1, sheet.max_row + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    
                    # Skip merged cells by checking if the cell is part of a merged range
                    is_merged = False
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            is_merged = True
                            break
                    
                    if is_merged:
                        continue
                        
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set adjusted width
                adjusted_width = min(max_length + 2, 25) if max_length > 0 else 12
                sheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            print(f"Error in auto-adjust columns: {e}")
            # Don't fail the whole export if column adjustment fails